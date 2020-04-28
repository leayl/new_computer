"""
使用openpyxl模块处理xlsx文件
可直接使用openpyxl操作xlsx文件，不必像操作xls文件一样需要xlutils模块作为中间转化
"""
import datetime
from copy import copy

import openpyxl
from openpyxl.styles import numbers
from openpyxl.cell import WriteOnlyCell


def read_xlsx(filepath):
    """
    使用openpyxl读取xlsx文件内容
    """
    # 获取工作簿
    wb = openpyxl.load_workbook(filepath)
    # 获取默认sheet
    ws = wb.active
    # 或者根据sheet名字获取sheet
    # ws = wb.get_sheet_by_name(sheet_name)
    # ws = wb.[sheet_name]

    # ws.rows或ws.columns获取表格的行或列的生成器
    for row in ws.rows:
        for cell in row:
            print(cell.value)

    # 获取某一单元格
    cell = ws["A1"]
    cell = ws.cell(1, 1)
    print(cell.value)

    # 获取多单元格
    # 1. 切片
    cells = ws["A1":"C2"]  # 获得每行单元格元祖组成的元祖,如：((1行1列，1行2列，1行3列),(2行1列，2行2列，2行3列))
    for row in cells:
        for cell in row:
            print(cell.value)
    # 2.指定行或列
    row1 = ws[1]  # col1 = ws["A:C"]，结果元祖以行为单位
    for cell in row1:
        print(cell.value)
    row1to2 = ws[1:2]  # col1to3 = ws["A:C"]，注意结果元祖以列为单位
    for row in row1to2:
        for cell in row:
            print(cell.value)


def create_xlsx(filepath):
    """
    使用openpyxl新建xlsx文件
    """
    # 创建工作簿
    wb = openpyxl.Workbook()
    ws = wb.create_sheet("sheet1", 0)  # 创建sheet，可指定位置
    ws2 = wb.create_sheet("sheet2", 1)

    datas = [
        ("标题1", "标题2", "标题3", "标题4"),
        ("内容11", "内容12", "内容13", "内容14"),
        ("内容21", "内容32", "内容23", "内容24"),
    ]

    # 插入行
    for row_data in datas:
        ws.append(row_data)

    # 直接修改单元格值
    ws.cell(2, 1, 'sss')

    wb.save(filepath)


def modify_xlsx(filepath):
    """
    使用openpyxl修改xlsx文件
    """
    wb = openpyxl.load_workbook(filepath)
    ws = wb["Sheet1"]
    new_sheet = wb.create_sheet("sheet_create")
    # wb.remove(new_sheet)  # 删除sheet
    for i in range(5):
        new_sheet.append(range(4))
    # 合并单元格,仅保留左上角单元格数据
    new_sheet.merge_cells("A1:B2")
    # 拆分单元格,数据填充在左上角单元格，其它单元格无数据
    new_sheet.unmerge_cells("A1:B2")

    nrows = ws.max_row  # 获取行数
    ncols = ws.max_column  # 获取列数
    for row in range(0, nrows):
        for col in range(0, ncols):
            cell = ws.cell(row + 1, col + 1)
            cell_value = cell.value
            if isinstance(cell_value, datetime.datetime):
                cell.value = cell_value + datetime.timedelta(days=1)

    wb.save(filepath)


def copy_temp_xlsx(filepath):
    """
    使用openpyxl复制单元格格式
    """
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    new_ws1 = wb.copy_worksheet(ws)  # 拷贝sheet，部分格式肯能无法拷贝
    new_ws1.title = "直接拷贝"

    wb.save(filepath)


num_format = numbers.FORMAT_NUMBER
pdf_format = numbers.FORMAT_PERCENTAGE_00


def set_row_data(cell_lis, ws, row, data_lis, string_lis):
    for index, data in enumerate(data_lis):
        col = index * 3 + 4
        cell_num = WriteOnlyCell(ws, data)
        cell_num.number_format = num_format  # 单元格值设置为数字
        # ws[f"{string_lis[col - 1]}{row}"].number_format = num_format  # 单元格值设置为数字
        # ws[f"{string_lis[col - 1]}{row}"].value = data  # 单元格值
        cell_pdf = WriteOnlyCell(ws, f"={string_lis[col - 1]}{row}/{string_lis[col - 1]}2")
        cell_pdf.number_format = pdf_format  # 当前所占百分比设置为百分数
        # ws[f"{string_lis[col]}{row}"].number_format = pdf_format  # 当前所占百分比设置为百分数
        # ws[f"{string_lis[col]}{row}"].value = f"={string_lis[col - 1]}{row}/{string_lis[col - 1]}2"  # 当前所占百分比
        # ws[f"{string_lis[col + 1]}{row}"].number_format = pdf_format  # 累计百分比设置为百分数
        if row == 3:
            cell_cdf = WriteOnlyCell(ws, f"={string_lis[col]}{row}")
            # ws[f"{string_lis[col + 1]}{row}"].value = f"={string_lis[col]}{row}"  # 第一条数据的累计百分比

        else:
            cell_cdf = WriteOnlyCell(ws, f"={string_lis[col]}{row}+{string_lis[col + 1]}{row - 1}")
            # ws[
            #     f"{string_lis[col + 1]}{row}"].value = f"={string_lis[col]}{row}+{string_lis[col + 1]}{row - 1}"  # 当前累计百分比
        cell_cdf.number_format = pdf_format  # 当前所占百分比设置为百分数
        cell_lis.append(cell_num)
        cell_lis.append(cell_pdf)
        cell_lis.append(cell_cdf)
    ws.append(cell_lis)


def create_pdf_xlsx():
    import string
    string_lis = list(string.ascii_uppercase)
    wb = openpyxl.Workbook(write_only=True)
    ws = wb.active
    header = ["序号", "callid", "enid", "参数一", "PDF", "CDF", "参数二", "PDF", "CDF", "参数三", "PDF", "CDF", "参数死四", "PDF",
              "CDF"]
    ws.append(header)
    ws.cell(2, 1).value = "汇总"
    ws.merge_cells("A2:C2")
    for i in range(4):
        col = i * 3 + 4
        ws[f"{string_lis[col - 1]}2"].number_format = num_format  # 总和格式
        ws[f"{string_lis[col - 1]}2"].value = 0  # 更新总和值
        ws[f"{string_lis[col]}2"].number_format = pdf_format  # 总和百分比
        ws[f"{string_lis[col]}2"].value = 1
        ws[f"{string_lis[col + 1]}2"].number_format = pdf_format  # 总累计百分比
    for i in range(2 ** 20):
        data = [i, i + 1, i + 2, i + 3]
        row = i + 3
        ws.cell(row, 2).value = f"callid_{i + 1}"
        ws.cell(row, 3).value = f"ENid_{i + 1}"
        set_row_data(ws, row, data, string_lis)
    nrow = ws.max_row
    for i in range(4):
        col = i * 3 + 4
        ws[f"{string_lis[col - 1]}2"].value = f"=SUM({string_lis[col - 1]}3:{string_lis[col - 1]}{nrow})"  # 更新总和值
    wb.save("excel_ouput/pdf_xlsx.xlsx")


def create_pdf_readonly_xlsx():
    import string
    string_lis = list(string.ascii_uppercase)
    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet()
    header = ["序号", "callid", "enid", "参数一", "PDF", "CDF", "参数二", "PDF", "CDF", "参数三", "PDF", "CDF", "参数死四", "PDF",
              "CDF"]
    ws.append(header)
    # ws.cell(2, 1).value = "汇总"
    # ws.merge_cells("A2:C2")
    row_1 = ["汇总", "", ""]
    for i in range(4):
        col = i * 3 + 4
        total_cell = WriteOnlyCell(ws, f"=SUM({string_lis[col - 1]}3:{string_lis[col - 1]}1048576)")
        total_cell.number_format = num_format  # 总和格式
        total_pdf_cell = WriteOnlyCell(ws, 1)
        total_pdf_cell.number_format = pdf_format  # 总和百分比
        total_cdf_cell = WriteOnlyCell(ws, 1)
        total_cdf_cell.number_format = pdf_format  # 总累计百分比
        row_1.append(total_cell)
        row_1.append(total_pdf_cell)
        row_1.append(total_cdf_cell)
    ws.append(row_1)
    for i in range(2 ** 20):
        data = [i, i + 1, i + 2, i + 3]
        row = i + 3
        cell_lis = ['', f"callid_{i + 1}", f"ENid_{i + 1}"]
        set_row_data(cell_lis, ws, row, data, string_lis)
    # nrow = ws.max_row
    # for i in range(4):
    #     col = i * 3 + 4
    #     ws[f"{string_lis[col - 1]}2"].value = f"=SUM({string_lis[col - 1]}3:{string_lis[col - 1]}{nrow})"  # 更新总和值
    wb.save("excel_ouput/pdf_xlsx.xlsx")


def read_only():
    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet()

    cell_total = WriteOnlyCell(ws, "=SUM(A2:A1048576)")
    cell_total.number_format = num_format
    cell_str = WriteOnlyCell(ws, "")
    ws.append([cell_total, cell_str])

    for i in range(2 ** 20 - 1):
        cell_cur = WriteOnlyCell(ws, i)
        cell_cur.number_format = num_format
        cell_pdf = WriteOnlyCell(ws, F"=A{i + 2}/A1")
        cell_pdf.number_format = pdf_format
        ws.append([cell_cur, cell_pdf])
    wb.save("ohmygod.xlsx")


if __name__ == '__main__':
    # filepath = "excel_ouput/temp_xlsx.xlsx"
    # new_path = "excel_ouput/new_xlsx.xlsx"
    # read_xlsx(filepath)
    # create_xlsx(new_path)
    # modify_xlsx(filepath)
    # copy_temp_xlsx(filepath)
    create_pdf_readonly_xlsx()
    # read_only()
